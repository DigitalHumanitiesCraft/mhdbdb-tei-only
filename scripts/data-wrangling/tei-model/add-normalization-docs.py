#!/usr/bin/env python3
"""
Issue #67: Document abbreviation/ligature resolutions in teiHeaders

Parses TEXT_DATA_TABLE.xlsx DESCRIPTION column for abbreviation info and
inserts <normalization> elements into <editorialDecl> of affected TEI files.

Patterns recognized:
  - "Übergeschriebene Vokale wurden aufgelöst"
  - "Ligatur von oe wurde aufgelöst"
  - "Hochgestellte o/e zu uo/ue aufgelöst"
  - "Diakritika reduziert"
  - Specific abbreviation descriptions

Usage:
    python scripts/data-wrangling/tei-model/add-normalization-docs.py [--dry-run]
"""

import argparse
import logging
import sys
from collections import defaultdict
from pathlib import Path

import openpyxl
from lxml import etree

logging.basicConfig(level=logging.INFO, format='%(levelname)s: %(message)s')
logger = logging.getLogger(__name__)

TEI_NS = 'http://www.tei-c.org/ns/1.0'
TEI = f'{{{TEI_NS}}}'

XLSX_FILE = Path('scripts/data-wrangling/tei-model/TEXT_DATA_TABLE.xlsx')
TEI_DIR = Path('tei')

ABBREV_KEYWORDS = [
    'abbreviat', 'ligatur', 'aufgelöst', 'aufgeloest',
    'diakrit', 'hochgestellt', 'übergeschrieb',
]


def extract_abbreviation_info():
    """Parse XLSX and return sigle → list of abbreviation descriptions."""
    wb = openpyxl.load_workbook(str(XLSX_FILE), read_only=True)
    ws = wb.active
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    desc_col = headers.index('DESCRIPTION')
    id_col = headers.index('ID')

    sigle_abbrevs = defaultdict(list)

    for row in ws.iter_rows(min_row=2, values_only=True):
        desc = str(row[desc_col]) if row[desc_col] else ''
        sigle = str(row[id_col]).strip() if row[id_col] else ''
        if not sigle:
            continue

        for line in desc.split('\n'):
            line = line.strip()
            if line and any(kw in line.lower() for kw in ABBREV_KEYWORDS):
                if line not in sigle_abbrevs[sigle]:
                    sigle_abbrevs[sigle].append(line)

    wb.close()
    return dict(sigle_abbrevs)


def add_normalization(tei_file, descriptions, dry_run):
    """Add <normalization> to <editorialDecl> in a TEI file."""
    parser = etree.XMLParser(remove_blank_text=False)
    tree = etree.parse(str(tei_file), parser)
    root = tree.getroot()

    editorial = root.find(f'.//{TEI}editorialDecl')
    if editorial is None:
        logger.warning(f'  No <editorialDecl> in {tei_file.name}')
        return False

    # Check if <normalization> already exists
    existing = editorial.find(f'{TEI}normalization')
    if existing is not None:
        logger.info(f'  {tei_file.name}: <normalization> already exists, skipping')
        return False

    if dry_run:
        return True

    # Build <normalization> element
    norm = etree.SubElement(editorial, f'{TEI}normalization')
    norm.text = '\n            '
    norm.tail = '\n        '

    for i, desc in enumerate(descriptions):
        p = etree.SubElement(norm, f'{TEI}p')
        p.text = desc
        p.tail = '\n            ' if i < len(descriptions) - 1 else '\n          '

    tree.write(str(tei_file), encoding='UTF-8', xml_declaration=True)
    return True


def main():
    parser = argparse.ArgumentParser(description='Issue #67: Add normalization docs to teiHeaders')
    parser.add_argument('--dry-run', action='store_true',
                        help='Show what would change without modifying files')
    args = parser.parse_args()

    if args.dry_run:
        logger.info('DRY RUN — no files will be modified')

    # Extract abbreviation info from XLSX
    logger.info('Reading TEXT_DATA_TABLE.xlsx...')
    abbrev_info = extract_abbreviation_info()
    logger.info(f'Found abbreviation info for {len(abbrev_info)} texts')

    # Process TEI files
    updated = 0
    skipped = 0
    missing = 0

    for sigle in sorted(abbrev_info.keys()):
        tei_file = TEI_DIR / f'{sigle}.tei.xml'
        if not tei_file.exists():
            missing += 1
            continue

        if add_normalization(tei_file, abbrev_info[sigle], args.dry_run):
            updated += 1
        else:
            skipped += 1

    logger.info('')
    logger.info(f'{"Would update" if args.dry_run else "Updated"}: {updated} files')
    logger.info(f'Skipped (already has <normalization>): {skipped}')
    if missing:
        logger.info(f'Missing TEI files: {missing}')

    return 0


if __name__ == '__main__':
    sys.exit(main())
