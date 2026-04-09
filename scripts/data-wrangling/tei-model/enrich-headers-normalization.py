#!/usr/bin/env python3
"""
Phase D1: Add <normalization> block from TEXT_DATA_TABLE.xlsx (Issue #67)

124 texts have DESCRIPTION entries with abbreviation/ligature documentation.
This script extracts the relevant text and adds it as:

  <editorialDecl>
    ...existing content...
    <normalization>
      <p>...description from XLSX...</p>
    </normalization>
  </editorialDecl>

Source: scripts/data-wrangling/tei-model/TEXT_DATA_TABLE.xlsx (DESCRIPTION column)
Match: ID column → TEI filename sigle

Scope: all *.tei.xml in tei/ that have matching DESCRIPTION entries

Usage:
    python scripts/data-wrangling/tei-model/enrich-headers-normalization.py [--dry-run]
"""

import argparse
import logging
import sys
from pathlib import Path
from lxml import etree

logging.basicConfig(level=logging.INFO, format='%(levelname)s: %(message)s')
logger = logging.getLogger(__name__)

TEI_NS = 'http://www.tei-c.org/ns/1.0'
TEI = f'{{{TEI_NS}}}'

TEI_DIR = Path('tei')
XLSX_PATH = Path('scripts/data-wrangling/tei-model/TEXT_DATA_TABLE.xlsx')


def load_descriptions() -> dict:
    """Load DESCRIPTION column from XLSX. Returns {sigle: description}."""
    try:
        import openpyxl
    except ImportError:
        logger.error('openpyxl required: pip install openpyxl')
        sys.exit(1)

    wb = openpyxl.load_workbook(str(XLSX_PATH), read_only=True)
    ws = wb.active
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    desc_idx = headers.index('DESCRIPTION')
    id_idx = headers.index('ID')

    descriptions = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        sigle = row[id_idx]
        desc = row[desc_idx]
        if sigle and desc and desc.strip():
            # Clean up: normalize whitespace, keep paragraph breaks
            paragraphs = [p.strip() for p in desc.strip().split('\n\n') if p.strip()]
            descriptions[sigle] = paragraphs

    wb.close()
    return descriptions


def migrate_file(tei_file: Path, paragraphs: list, dry_run: bool) -> bool:
    """Add <normalization> to one file. Returns True if changed."""
    parser = etree.XMLParser(remove_blank_text=False)
    tree = etree.parse(str(tei_file), parser)
    root = tree.getroot()

    # Check if normalization already exists
    existing = root.find(f'.//{TEI}normalization')
    if existing is not None:
        return False

    # Find editorialDecl
    editorial_decl = root.find(f'.//{TEI}editorialDecl')
    if editorial_decl is None:
        logger.warning(f'  No <editorialDecl> in {tei_file.name}')
        return False

    if dry_run:
        return True

    # Build <normalization> with <p> children
    normalization = etree.SubElement(editorial_decl, f'{TEI}normalization')
    for para_text in paragraphs:
        p = etree.SubElement(normalization, f'{TEI}p')
        p.text = para_text

    tree.write(str(tei_file), encoding='UTF-8', xml_declaration=True)
    return True


def main():
    parser = argparse.ArgumentParser(
        description='Phase D1: Add normalization info from TEXT_DATA_TABLE.xlsx')
    parser.add_argument('--dry-run', action='store_true',
                        help='Show what would change without modifying files')
    args = parser.parse_args()

    if args.dry_run:
        logger.info('DRY RUN — no files will be modified')

    descriptions = load_descriptions()
    logger.info(f'Loaded {len(descriptions)} descriptions from {XLSX_PATH.name}')

    added = 0
    skipped = 0
    missing = 0

    for sigle, paragraphs in sorted(descriptions.items()):
        tei_file = TEI_DIR / f'{sigle}.tei.xml'
        if not tei_file.exists():
            logger.debug(f'{sigle}: TEI file not found')
            missing += 1
            continue

        if migrate_file(tei_file, paragraphs, args.dry_run):
            added += 1
        else:
            skipped += 1

    logger.info('')
    logger.info(f'{"Would add" if args.dry_run else "Added"} <normalization> to {added} files')
    if skipped:
        logger.info(f'Skipped {skipped} (already had <normalization>)')
    if missing:
        logger.info(f'Missing TEI files: {missing}')
    return 0


if __name__ == '__main__':
    sys.exit(main())
